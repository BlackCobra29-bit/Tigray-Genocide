import base64
import csv
import re
from functools import lru_cache
from io import BytesIO
from tempfile import SpooledTemporaryFile
from textwrap import wrap as text_wrap

from django.conf import settings
from django.http import FileResponse, StreamingHttpResponse
from django.utils import timezone

from .civilian_management import get_filtered_ordered_civilian_queryset


EXPORT_HEADERS = (
    "Index",
    "Full Name",
    "Gender",
    "Age",
    "Atrocity",
    "Place of Killing",
    "Woreda",
    "Source",
    "Date of Event",
    "Remark",
)
SUPPORTED_EXPORT_FORMATS = {"csv", "xlsx", "pdf"}
EXPORT_ITERATOR_CHUNK_SIZE = 500


def _export_filename(extension):
    date_stamp = timezone.localdate().isoformat()
    return f"verified-civilian-victims-{date_stamp}.{extension}"


def _plain_value(value, fallback="Unknown"):
    return str(value) if value not in (None, "") else fallback


def _iter_export_rows(queryset):
    for index, victim in enumerate(
        queryset.iterator(chunk_size=EXPORT_ITERATOR_CHUNK_SIZE),
        start=1,
    ):
        yield (
            index,
            victim.full_name.title(),
            victim.gender,
            _plain_value(victim.age),
            victim.perpetrator,
            _plain_value(victim.place_of_killing),
            _plain_value(victim.woreda),
            _plain_value(victim.source),
            (
                victim.date_of_event.strftime("%d-%b-%Y")
                if victim.date_of_event
                else "Undated"
            ),
            _plain_value(victim.remark),
        )


def build_civilian_export_payload(user, params):
    queryset = get_filtered_ordered_civilian_queryset(user, params)
    rows = [list(row) for row in _iter_export_rows(queryset)]
    return {
        "recordsFiltered": len(rows),
        "data": rows,
    }


def _safe_spreadsheet_value(value):
    value = str(value)
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


class _CsvEcho:
    def write(self, value):
        return value


def _build_csv_response(queryset):
    writer = csv.writer(_CsvEcho())

    def rows():
        yield "\ufeff"
        yield writer.writerow(EXPORT_HEADERS)
        for row in _iter_export_rows(queryset):
            yield writer.writerow(
                [_safe_spreadsheet_value(value) for value in row]
            )

    response = StreamingHttpResponse(
        rows(),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_export_filename("csv")}"'
    )
    return response


def _build_xlsx_response(queryset):
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    output = SpooledTemporaryFile(max_size=5 * 1024 * 1024)
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Verified Civilian Victims")
    worksheet.freeze_panes = "A2"
    column_widths = (10, 34, 12, 10, 38, 30, 24, 30, 18, 45)
    for column_index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    header_fill = PatternFill("solid", fgColor="2A3F54")
    header_font = Font(color="FFFFFF", bold=True)
    header_cells = []
    for heading in EXPORT_HEADERS:
        cell = WriteOnlyCell(worksheet, value=heading)
        cell.fill = header_fill
        cell.font = header_font
        header_cells.append(cell)
    worksheet.append(header_cells)

    for row in _iter_export_rows(queryset):
        worksheet.append(
            [_safe_spreadsheet_value(value) for value in row]
        )

    workbook.save(output)
    output.seek(0)
    return FileResponse(
        output,
        as_attachment=True,
        filename=_export_filename("xlsx"),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


@lru_cache(maxsize=1)
def _get_pdf_font_name():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    vfs_path = (
        settings.BASE_DIR
        / "static"
        / "admlte"
        / "datatable"
        / "vfs_fonts.js"
    )
    try:
        vfs_text = vfs_path.read_text(encoding="utf-8")
        encoded_font = re.search(
            r'"nyala\.ttf"\s*:\s*"([^"]+)"',
            vfs_text,
        )
        if encoded_font is None:
            return "Helvetica"
        font_data = BytesIO(base64.b64decode(encoded_font.group(1)))
        pdfmetrics.registerFont(TTFont("CivilianExport", font_data))
        return "CivilianExport"
    except (OSError, ValueError):
        return "Helvetica"


def _build_pdf_response(queryset):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import TABLOID, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfgen import canvas

    font_name = _get_pdf_font_name()
    rows = list(_iter_export_rows(queryset))
    output = SpooledTemporaryFile(max_size=5 * 1024 * 1024)
    page_size = landscape(TABLOID)
    pdf = canvas.Canvas(
        output,
        pagesize=page_size,
        pageCompression=1,
    )
    pdf.setTitle("Verified Civilian Victims")
    pdf.setAuthor("tigraygenocide.com")

    page_width, page_height = page_size
    column_widths = tuple(
        width * mm
        for width in (14, 48, 22, 14, 55, 45, 36, 42, 28, 69)
    )
    table_width = sum(column_widths)
    table_x = (page_width - table_width) / 2
    top_margin = 12 * mm
    bottom_margin = 12 * mm
    header_height = 7 * mm
    font_size = 5.5
    line_height = 6.5
    cell_padding = 2
    page_number = 0

    def wrap_cell(value, width):
        text = str(value)
        average_character_width = max(
            pdfmetrics.stringWidth(
                "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz",
                font_name,
                font_size,
            ) / 52,
            1,
        )
        maximum_characters = max(
            int((width - (2 * cell_padding)) / average_character_width),
            1,
        )
        lines = []
        for source_line in text.splitlines() or [""]:
            lines.extend(
                text_wrap(
                    source_line,
                    width=maximum_characters,
                    break_long_words=True,
                    break_on_hyphens=False,
                )
                or [""]
            )
        return lines

    def decorate_page():
        pdf.saveState()
        pdf.setFont(font_name, 34)
        pdf.setFillColor(colors.HexColor("#E8E8E8"))
        pdf.translate(page_width / 2, page_height / 2)
        pdf.rotate(30)
        pdf.drawCentredString(0, 0, "© tigraygenocide.com")
        pdf.rotate(-30)
        pdf.translate(-page_width / 2, -page_height / 2)
        pdf.setFont(font_name, 8)
        pdf.setFillColor(colors.HexColor("#4F6B85"))
        pdf.drawString(10 * mm, 7 * mm, "Visit tigraygenocide.com")
        pdf.drawRightString(
            page_width - 10 * mm,
            7 * mm,
            f"Page {page_number}",
        )
        pdf.restoreState()

    def draw_header(top):
        pdf.saveState()
        pdf.setFillColor(colors.HexColor("#2A3F54"))
        pdf.setStrokeColor(colors.HexColor("#B8C2CC"))
        pdf.rect(
            table_x,
            top - header_height,
            table_width,
            header_height,
            stroke=1,
            fill=1,
        )
        pdf.setFillColor(colors.white)
        pdf.setFont(font_name, 6.5)
        x_position = table_x
        for heading, width in zip(EXPORT_HEADERS, column_widths):
            pdf.line(
                x_position,
                top - header_height,
                x_position,
                top,
            )
            pdf.drawCentredString(
                x_position + (width / 2),
                top - header_height + 6,
                heading,
            )
            x_position += width
        pdf.restoreState()
        return top - header_height

    def start_page(first_page=False):
        nonlocal page_number
        page_number += 1
        decorate_page()
        y_position = page_height - top_margin
        if first_page:
            pdf.setFillColor(colors.HexColor("#2A3F54"))
            pdf.setFont(font_name, 18)
            pdf.drawCentredString(
                page_width / 2,
                y_position,
                "Verified Civilian Victims",
            )
            y_position -= 8 * mm
            pdf.setFillColor(colors.black)
            pdf.setFont(font_name, 9)
            pdf.drawCentredString(
                page_width / 2,
                y_position,
                "A compilation of the verified list of civilian victims "
                "from different sources. "
                f"Total number of victims: {len(rows):,}",
            )
            y_position -= 7 * mm
        return draw_header(y_position)

    y_position = start_page(first_page=True)
    full_page_line_capacity = int(
        (
            page_height
            - (2 * top_margin)
            - bottom_margin
            - header_height
            - (2 * cell_padding)
        )
        / line_height
    )

    for row_number, row in enumerate(rows, start=1):
        wrapped_cells = [
            wrap_cell(value, width)
            for value, width in zip(row, column_widths)
        ]

        while any(wrapped_cells):
            maximum_lines = max(len(lines) for lines in wrapped_cells)
            available_lines = int(
                (
                    y_position
                    - bottom_margin
                    - (2 * cell_padding)
                )
                / line_height
            )
            if (
                maximum_lines <= full_page_line_capacity
                and maximum_lines > available_lines
            ) or available_lines < 1:
                pdf.showPage()
                y_position = start_page()
                available_lines = int(
                    (
                        y_position
                        - bottom_margin
                        - (2 * cell_padding)
                    )
                    / line_height
                )

            lines_in_segment = min(maximum_lines, available_lines)
            segment_height = (
                lines_in_segment * line_height
                + (2 * cell_padding)
            )
            row_bottom = y_position - segment_height

            pdf.saveState()
            pdf.setFillColor(
                colors.white
                if row_number % 2
                else colors.HexColor("#F4F6F9")
            )
            pdf.setStrokeColor(colors.HexColor("#B8C2CC"))
            pdf.rect(
                table_x,
                row_bottom,
                table_width,
                segment_height,
                stroke=1,
                fill=1,
            )
            pdf.setFillColor(colors.black)
            pdf.setFont(font_name, font_size)

            x_position = table_x
            next_wrapped_cells = []
            for lines, width in zip(wrapped_cells, column_widths):
                pdf.line(x_position, row_bottom, x_position, y_position)
                segment_lines = lines[:lines_in_segment]
                remaining_lines = lines[lines_in_segment:]
                next_wrapped_cells.append(remaining_lines)

                text_object = pdf.beginText(
                    x_position + cell_padding,
                    y_position - cell_padding - font_size,
                )
                text_object.setFont(font_name, font_size)
                text_object.setLeading(line_height)
                for line in segment_lines:
                    text_object.textLine(line)
                pdf.drawText(text_object)
                x_position += width

            pdf.restoreState()
            y_position = row_bottom
            wrapped_cells = next_wrapped_cells

            if any(wrapped_cells):
                pdf.showPage()
                y_position = start_page()

    pdf.save()
    output.seek(0)
    return FileResponse(
        output,
        as_attachment=True,
        filename=_export_filename("pdf"),
        content_type="application/pdf",
    )


def build_civilian_export_response(export_format, user, params):
    queryset = get_filtered_ordered_civilian_queryset(user, params)
    if export_format == "csv":
        return _build_csv_response(queryset)
    if export_format == "xlsx":
        return _build_xlsx_response(queryset)
    if export_format == "pdf":
        return _build_pdf_response(queryset)
    raise ValueError(f"Unsupported export format: {export_format}")
