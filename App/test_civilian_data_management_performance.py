from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import Resolver404, resolve, reverse
from openpyxl import load_workbook

from .models import Administrator, Civilian_victims, Tigray_woreda
from .templatetags.custom import _extract_urls, extract_and_join_urls


class CivilianDataManagementPerformanceTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        user_model = get_user_model()
        cls.superuser = user_model.objects.create_user(
            username="civilian-superuser",
            password="dashboard-password",
            is_staff=True,
            is_superuser=True,
        )
        cls.administrator = user_model.objects.create_user(
            username="civilian-administrator",
            password="dashboard-password",
            is_staff=True,
        )
        cls.denied_user = user_model.objects.create_user(
            username="civilian-denied",
            password="dashboard-password",
            is_staff=True,
        )
        Administrator.objects.create(user=cls.superuser)
        Administrator.objects.create(
            user=cls.administrator,
            civilian_role=True,
        )
        Administrator.objects.create(user=cls.denied_user)

        cls.woreda = Tigray_woreda.objects.create(
            woreda_name="Management Woreda",
            latitude="13.5",
            longitude="39.5",
            zone="Central Tigray",
        )
        records = [
            Civilian_victims(
                author=cls.superuser,
                full_name=f"Managed Victim {index:02}",
                gender="Female" if index % 2 else "Male",
                age=20 + index,
                place_of_killing="Management Place",
                woreda=cls.woreda,
                zone="Central Tigray",
                source="Management Source",
                source_link=f"https://example.com/source/{index}",
                perpetrator="Killed by Eritrean forces",
                remark="Management remark",
                approval=True,
            )
            for index in range(30)
        ]
        records.extend(
            [
                Civilian_victims(
                    author=cls.administrator,
                    full_name="Administrator Approved Victim",
                    gender="Female",
                    woreda=cls.woreda,
                    zone="Central Tigray",
                    source="Administrator Source",
                    source_link="https://example.com/administrator-approved",
                    perpetrator="Killed by Ethiopian forces",
                    approval=True,
                ),
                Civilian_victims(
                    author=cls.administrator,
                    full_name="Administrator Pending Victim",
                    gender="Male",
                    woreda=cls.woreda,
                    zone="Central Tigray",
                    source="Administrator Source",
                    source_link="https://example.com/administrator-pending",
                    perpetrator="Killed by Ethiopian forces",
                    approval=False,
                ),
            ]
        )
        Civilian_victims.objects.bulk_create(records)

    def setUp(self):
        cache.clear()

    def test_page_renders_only_the_server_side_table_shell(self):
        self.client.force_login(self.superuser)

        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(reverse("civilian-data-management"))

        html = response.content.decode()
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="civilian-data-management-table"')
        self.assertContains(response, "serverSide: true")
        self.assertContains(response, reverse("civilian-data-management-data"))
        self.assertNotContains(response, "Managed Victim 00")
        self.assertIn("<tbody></tbody>", html)
        self.assertLess(len(response.content), 100_000)
        self.assertLessEqual(len(queries), 5)

    def test_page_loads_only_lean_datatables_assets(self):
        self.client.force_login(self.superuser)
        response = self.client.get(reverse("civilian-data-management"))
        html = response.content.decode()

        self.assertNotIn("/static/froala_editor/", html)
        self.assertNotIn("/static/admin/js/vendor/jquery/jquery.js", html)
        self.assertNotIn("/static/admin_static/js/parsley.min.js", html)
        self.assertNotIn("/static/js/select.js", html)
        self.assertNotIn(
            '<script src="/static/admlte/datatable/'
            'dataTables.buttons.min.js"',
            html,
        )
        self.assertNotIn(
            '<script src="/static/admlte/datatable/pdfmake.min.js"',
            html,
        )
        self.assertNotIn(
            '<script src="/static/admlte/datatable/vfs_fonts.js"',
            html,
        )
        self.assertEqual(
            html.count("/static/admin_static/datatable/"),
            6,
        )

    def test_superuser_sees_export_buttons_without_heavy_export_libraries(self):
        self.client.force_login(self.superuser)
        response = self.client.get(reverse("civilian-data-management"))
        html = response.content.decode()

        self.assertContains(response, "Export current results:")
        self.assertContains(
            response,
            reverse("civilian-data-management-export", args=["pdf"]),
        )
        self.assertContains(
            response,
            reverse("civilian-data-management-export", args=["csv"]),
        )
        self.assertContains(
            response,
            reverse("civilian-data-management-export", args=["xlsx"]),
        )
        self.assertNotIn(
            '<script src="/static/admlte/datatable/'
            'dataTables.buttons.min.js"',
            html,
        )
        self.assertNotIn(
            '<script src="/static/admlte/datatable/buttons.html5.min.js"',
            html,
        )
        self.assertNotIn(
            '<script src="/static/admlte/datatable/pdfmake.min.js"',
            html,
        )
        self.assertNotIn(
            '<script src="/static/admlte/datatable/vfs_fonts.js"',
            html,
        )
        self.assertNotIn(
            '<script src="/static/admlte/datatable/jszip.min.js"',
            html,
        )
        self.assertContains(response, "legacyExportScriptUrls")
        self.assertContains(response, "pageSize: 'TABLOID'")
        self.assertContains(response, "orientation: 'landscape'")
        self.assertContains(response, "fillColor: '#23ffee'")
        self.assertContains(response, "doc.watermark")
        self.assertContains(response, "doc.defaultStyle.font = 'nyala'")

    def test_legacy_export_data_uses_the_current_filter_and_one_data_query(self):
        self.client.force_login(self.superuser)

        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(
                reverse("civilian-data-management-export-data"),
                {"search[value]": "Managed Victim 17"},
            )

        payload = response.json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["recordsFiltered"], 1)
        self.assertEqual(len(payload["data"]), 1)
        self.assertEqual(
            payload["data"][0][0:3],
            [1, "Managed Victim 17", "Female"],
        )
        self.assertLessEqual(len(queries), 3)

    def test_non_superuser_does_not_see_or_access_exports(self):
        self.client.force_login(self.administrator)
        page_response = self.client.get(reverse("civilian-data-management"))
        export_response = self.client.get(
            reverse("civilian-data-management-export", args=["csv"])
        )
        legacy_data_response = self.client.get(
            reverse("civilian-data-management-export-data")
        )

        self.assertNotContains(page_response, "Export current results:")
        self.assertEqual(export_response.status_code, 403)
        self.assertEqual(legacy_data_response.status_code, 403)

    def test_csv_export_contains_all_currently_filtered_rows(self):
        self.client.force_login(self.superuser)

        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(
                reverse("civilian-data-management-export", args=["csv"]),
                {
                    "search[value]": "Managed Victim 17",
                    "order[0][column]": 1,
                    "order[0][dir]": "asc",
                },
            )
            content = b"".join(response.streaming_content).decode("utf-8-sig")

        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        self.assertIn(
            "verified-civilian-victims-",
            response["Content-Disposition"],
        )
        self.assertIn("Managed Victim 17", content)
        self.assertNotIn("Managed Victim 16", content)
        self.assertLessEqual(len(queries), 3)

    def test_excel_export_is_valid_and_preserves_the_current_filter(self):
        self.client.force_login(self.superuser)
        response = self.client.get(
            reverse("civilian-data-management-export", args=["xlsx"]),
            {"search[value]": "Managed Victim 17"},
        )
        content = b"".join(response.streaming_content)
        workbook = load_workbook(BytesIO(content), read_only=True)
        worksheet = workbook["Verified Civilian Victims"]
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(rows[0][0:3], ("Index", "Full Name", "Gender"))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1], "Managed Victim 17")

    def test_pdf_export_is_a_valid_filtered_download(self):
        self.client.force_login(self.superuser)
        response = self.client.get(
            reverse("civilian-data-management-export", args=["pdf"]),
            {"search[value]": "Managed Victim 17"},
        )
        content = b"".join(response.streaming_content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(content.startswith(b"%PDF-"))
        self.assertGreater(len(content), 1_000)

    def test_invalid_export_format_returns_not_found(self):
        self.client.force_login(self.superuser)
        response = self.client.get(
            reverse("civilian-data-management-export", args=["invalid"])
        )

        self.assertEqual(response.status_code, 404)

    def test_old_export_page_and_route_are_removed(self):
        self.client.force_login(self.superuser)

        with self.assertRaises(Resolver404):
            resolve("/Civilian-victims/")

        page_response = self.client.get(reverse("civilian-data-management"))
        self.assertNotContains(page_response, "/Civilian-victims/")

    def test_superuser_endpoint_is_paginated_and_has_no_woreda_n_plus_one(self):
        self.client.force_login(self.superuser)
        params = {
            "draw": 3,
            "start": 0,
            "length": 10,
        }

        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(
                reverse("civilian-data-management-data"),
                params,
            )

        payload = response.json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["draw"], 3)
        self.assertEqual(payload["recordsTotal"], 31)
        self.assertEqual(payload["recordsFiltered"], 31)
        self.assertEqual(len(payload["data"]), 10)
        self.assertTrue(
            all("Management Woreda" in row[6] for row in payload["data"])
        )
        self.assertLess(len(response.content), 100_000)
        self.assertLessEqual(len(queries), 5)

    def test_endpoint_preserves_search_and_sorting(self):
        self.client.force_login(self.superuser)
        response = self.client.get(
            reverse("civilian-data-management-data"),
            {
                "draw": 1,
                "start": 0,
                "length": 10,
                "search[value]": "Managed Victim 17",
                "order[0][column]": 1,
                "order[0][dir]": "desc",
            },
        )

        payload = response.json()
        self.assertEqual(payload["recordsFiltered"], 1)
        self.assertEqual(len(payload["data"]), 1)
        self.assertEqual(payload["data"][0][1], "Managed Victim 17")
        self.assertIn(
            '<a href="https://example.com/source/17"',
            payload["data"][0][7],
        )

    def test_administrator_endpoint_is_limited_to_authored_records(self):
        self.client.force_login(self.administrator)
        response = self.client.get(
            reverse("civilian-data-management-data"),
            {"draw": 1, "start": 0, "length": 10},
        )

        payload = response.json()
        names = {row[2] for row in payload["data"]}
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["recordsTotal"], 2)
        self.assertEqual(
            names,
            {
                "Administrator Approved Victim",
                "Administrator Pending Victim",
            },
        )
        self.assertTrue(any("Approved" in row[1] for row in payload["data"]))
        self.assertTrue(any("Pending" in row[1] for row in payload["data"]))

    def test_endpoint_rejects_administrator_without_civilian_role(self):
        self.client.force_login(self.denied_user)
        response = self.client.get(
            reverse("civilian-data-management-data"),
            {"draw": 1, "start": 0, "length": 10},
        )

        self.assertEqual(response.status_code, 403)

    def test_url_extractor_reuses_cached_results(self):
        _extract_urls.cache_clear()
        value = "See https://example.com/repeated-source"

        first = extract_and_join_urls(value)
        second = extract_and_join_urls(value)

        self.assertEqual(first, second)
        self.assertEqual(_extract_urls.cache_info().hits, 1)
